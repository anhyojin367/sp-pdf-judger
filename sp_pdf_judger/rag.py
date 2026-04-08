from __future__ import annotations

import json
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import linear_kernel

from .config import UCUM_JSON_CANDIDATES, UCUM_JSONL_CANDIDATES, UCUM_XLSX_CANDIDATES
from .utils import clean_text


@dataclass
class RagDoc:
    idx: int
    text: str
    meta: dict


class UcumRagStore:
    def __init__(self) -> None:
        self.docs: list[RagDoc] = []
        self.word_vectorizer: TfidfVectorizer | None = None
        self.char_vectorizer: TfidfVectorizer | None = None
        self.word_matrix = None
        self.char_matrix = None
        self.loaded_sources: list[str] = []
        self._load()

    def _existing_paths(self, candidates: list[Path]) -> list[Path]:
        out: list[Path] = []
        seen: set[Path] = set()
        for path in candidates:
            if not path.exists():
                continue
            resolved = path.resolve()
            if resolved in seen:
                continue
            seen.add(resolved)
            out.append(path)
        return out

    def _normalize_for_search(self, text: str | None) -> str:
        text = clean_text(text)
        if not text:
            return ""

        replacements = {
            "µ": "u",
            "μ": "u",
            "Å": "Ao",
            "Å": "Ao",
            "℃": "°C",
            "㎍": "ug",
            "㎎": "mg",
            "㎖": "mL",
            "㎕": "uL",
            "㎛": "um",
            "·": ".",
            "⋅": ".",
            "∙": ".",
            "⁻": "-",
            "−": "-",
            "²": "2",
            "³": "3",
            "¹": "1",
        }
        for src, dst in replacements.items():
            text = text.replace(src, dst)

        text = text.replace("mmHg", "mm[Hg] mmHg")
        text = text.replace("µL", "uL µL")
        text = text.replace("µg", "ug µg")
        text = text.replace("µm", "um µm")
        text = text.replace("µmol", "umol µmol")
        text = text.replace("µS/cm", "uS/cm µS/cm")
        return clean_text(text)

    def _build_json_doc_text(self, item: dict) -> str:
        pieces = [
            item.get("kind"),
            item.get("code"),
            item.get("name"),
            item.get("symbol"),
            item.get("property"),
            item.get("ref_unit"),
            item.get("ref_value"),
            item.get("text"),
        ]
        return self._normalize_for_search(" ".join(str(x) for x in pieces if x not in (None, "")))

    def _build_jsonl_doc_text(self, item: dict) -> str:
        fields = [
            item.get("major_category_ko"),
            item.get("property"),
            item.get("display_name_ko"),
            item.get("display_name_en"),
            item.get("canonical_ucum"),
            item.get("canonical_ascii"),
            item.get("canonical_symbol"),
            item.get("display_symbol_ui"),
            " ".join(item.get("display_symbol_variants", []) or []),
            " ".join(item.get("aliases", []) or []),
            item.get("text"),
        ]
        return self._normalize_for_search(" ".join(str(x) for x in fields if x not in (None, "")))

    def _load_json_file(self, path: Path, docs: list[RagDoc], seen: set[tuple[str, str]]) -> None:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)

        for item in data:
            text = self._build_json_doc_text(item)
            if not text:
                continue
            key = (str(path.resolve()), item.get("code") or item.get("text") or text)
            if key in seen:
                continue
            seen.add(key)
            meta = dict(item)
            meta["rag_source_file"] = path.name
            meta["rag_source_type"] = "json"
            docs.append(RagDoc(idx=len(docs), text=text, meta=meta))

        if path.name not in self.loaded_sources:
            self.loaded_sources.append(path.name)

    def _load_jsonl_file(self, path: Path, docs: list[RagDoc], seen: set[tuple[str, str]]) -> None:
        with open(path, "r", encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                item = json.loads(line)
                text = self._build_jsonl_doc_text(item)
                if not text:
                    continue
                key = ("jsonl", item.get("canonical_ucum") or item.get("id") or text)
                if key in seen:
                    continue
                seen.add(key)
                meta = dict(item)
                meta["rag_source_file"] = path.name
                meta["rag_source_type"] = "jsonl"
                docs.append(RagDoc(idx=len(docs), text=text, meta=meta))

        if path.name not in self.loaded_sources:
            self.loaded_sources.append(path.name)

    def _load_xlsx_file(self, path: Path, docs: list[RagDoc], seen: set[tuple[str, str]]) -> None:
        wb = load_workbook(path, data_only=True)
        if not wb.sheetnames:
            return
        ws = wb[wb.sheetnames[0]]

        for row in ws.iter_rows(min_row=3, values_only=True):
            (
                row_no,
                ucum_code,
                description,
                comment,
                last_updated,
                version_correction,
                corrected_by,
                previous_row_no,
                previous_ucum_version,
                change_description,
                *_,
            ) = row + (None,) * max(0, 10 - len(row))

            if not ucum_code:
                continue

            text = self._normalize_for_search(
                " ".join(
                    str(x)
                    for x in [
                        ucum_code,
                        description,
                        comment,
                        previous_ucum_version,
                        change_description,
                        version_correction,
                    ]
                    if x not in (None, "")
                )
            )
            if not text:
                continue

            key = ("xlsx", str(ucum_code))
            if key in seen:
                continue
            seen.add(key)
            meta = {
                "row_no": row_no,
                "ucum_code": ucum_code,
                "description": description,
                "comment": comment,
                "last_updated": str(last_updated) if last_updated else None,
                "version_correction": version_correction,
                "corrected_by": corrected_by,
                "previous_row_no": previous_row_no,
                "previous_ucum_version": previous_ucum_version,
                "change_description": change_description,
                "rag_source_file": path.name,
                "rag_source_type": "xlsx",
            }
            docs.append(RagDoc(idx=len(docs), text=text, meta=meta))

        if path.name not in self.loaded_sources:
            self.loaded_sources.append(path.name)

    def _load(self) -> None:
        docs: list[RagDoc] = []
        seen: set[tuple[str, str]] = set()

        for path in self._existing_paths(UCUM_JSON_CANDIDATES):
            self._load_json_file(path, docs, seen)

        for path in self._existing_paths(UCUM_JSONL_CANDIDATES):
            self._load_jsonl_file(path, docs, seen)

        for path in self._existing_paths(UCUM_XLSX_CANDIDATES):
            self._load_xlsx_file(path, docs, seen)

        if not docs:
            raise FileNotFoundError(
                "UCUM RAG 소스를 찾을 수 없습니다. ucum_rag_docs.json / ucum_rag_image_units_ui_micro.jsonl / TableOfExampleUcumCodesForElectronicMessaging.xlsx 중 하나 이상이 필요합니다."
            )

        self.docs = docs
        corpus = [d.text for d in docs]

        self.word_vectorizer = TfidfVectorizer(
            ngram_range=(1, 2),
            lowercase=False,
            token_pattern=r"(?u)[\w\[\]/.%'*-]+",
        )
        self.char_vectorizer = TfidfVectorizer(
            analyzer="char_wb",
            ngram_range=(2, 5),
            lowercase=False,
        )
        self.word_matrix = self.word_vectorizer.fit_transform(corpus)
        self.char_matrix = self.char_vectorizer.fit_transform(corpus)

    def search(self, query: str, top_k: int = 5) -> list[RagDoc]:
        query = self._normalize_for_search(query)
        if (
            not query
            or not self.docs
            or self.word_vectorizer is None
            or self.char_vectorizer is None
            or self.word_matrix is None
            or self.char_matrix is None
        ):
            return []

        word_query = self.word_vectorizer.transform([query])
        char_query = self.char_vectorizer.transform([query])
        word_scores = linear_kernel(word_query, self.word_matrix).flatten()
        char_scores = linear_kernel(char_query, self.char_matrix).flatten()
        scores = (0.65 * word_scores) + (0.35 * char_scores)

        ranked = sorted(range(len(scores)), key=lambda i: scores[i], reverse=True)
        out: list[RagDoc] = []
        for i in ranked[:top_k]:
            if scores[i] <= 0:
                continue
            out.append(self.docs[i])
        return out